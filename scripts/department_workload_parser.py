"""Parse department workload Excel files into normalized draft CSVs."""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

SKIP_SHEET_PREFIXES = ("сводная таблица", "свод")
SKIP_SHEET_SUFFIXES = (" (копия)",)

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "discipline": ("наименование учебной дисциплины",),
    "groups_raw": ("шифр учебной группы",),
    "course": ("курс",),
    "group_count": ("количество групп",),
    "student_count": ("количество студентов",),
    "weeks": ("количество недель",),
    "semester_dates": ("начало и окончание занятий",),
    "lab_hours": ("количество часов на лабораторные",),
    "lab_title": ("название лр",),
    "equipment": ("лабораторное оборудование", "наименование стенда"),
    "capacity": ("количество рабочих мест",),
    "duration_minutes": ("время выполнения лр", "среднее время выполнения лр"),
    "room": ("аудитория №", "аудитория"),
    "training_center": ("уч №",),
    "teacher": ("преподавателя",),
    "teacher_requires_presence": ("лр проводимы только с преподавателем",),
    "comments": ("комментарии", "примечание"),
}


@dataclass(frozen=True)
class DepartmentProfile:
    key: str
    department_code: str
    department_title: str
    faculty_code: str
    faculty_title: str
    default_training_center: str
    discipline_code_prefix: str
    output_dir: str
    notes: str = ""
    filename_hints: tuple[str, ...] = ()


DEPARTMENTS: dict[str, DepartmentProfile] = {
    "met": DepartmentProfile(
        key="met",
        department_code="МЕТ",
        department_title="Кафедра металлургии",
        faculty_code="ФПМС",
        faculty_title="Факультет переработки минерального сырья",
        default_training_center="1",
        discipline_code_prefix="MET",
        output_dir="docs/csv_templates/metallurgy_draft",
        notes="Аудитория 3432 — КУЛ ФПМС (УЦ №1).",
        filename_hints=("металлургии", "кафедра 23"),
    ),
    "otf": DepartmentProfile(
        key="otf",
        department_code="ОТФ",
        department_title="Кафедра общей и технической физики",
        faculty_code="ИБИО",
        faculty_title="Институт базового инженерного образования",
        default_training_center="3",
        discipline_code_prefix="OTF",
        output_dir="docs/csv_templates/otf_draft",
        notes="Все ЛР — УЦ №3 (инженерный корпус).",
        filename_hints=("общей и технической физики", "кафедра 19", "(отф)"),
    ),
    "htpe": DepartmentProfile(
        key="htpe",
        department_code="ХТПЭ",
        department_title="Кафедра химических технологий и переработки энергоносителей",
        faculty_code="ФПМС",
        faculty_title="Факультет переработки минерального сырья",
        default_training_center="1",
        discipline_code_prefix="HTP",
        output_dir="docs/csv_templates/htpe_draft",
        filename_hints=("химических технологий", "кафедра 22", "хт и пэ"),
    ),
    "ofh": DepartmentProfile(
        key="ofh",
        department_code="ОФХ",
        department_title="Кафедра общей и физической химии",
        faculty_code="ФПМС",
        faculty_title="Факультет переработки минерального сырья",
        default_training_center="3",
        discipline_code_prefix="OFH",
        output_dir="docs/csv_templates/ofh_draft",
        notes="Лаборатории общей химии — УЦ №3 (инженерный корпус, studlab).",
        filename_hints=("физической химии", "кафедра 24", "(офх)"),
    ),
    "opi": DepartmentProfile(
        key="opi",
        department_code="ОПИ",
        department_title="Кафедра обогащения полезных ископаемых",
        faculty_code="ФПМС",
        faculty_title="Факультет переработки минерального сырья",
        default_training_center="1",
        discipline_code_prefix="OPI",
        output_dir="docs/csv_templates/opi_draft",
        filename_hints=("обогащения полезных", "кафедра 25", "(опи)"),
    ),
    "bp": DepartmentProfile(
        key="bp",
        department_code="БП",
        department_title="Кафедра безопасности производств",
        faculty_code="ГФ",
        faculty_title="Горный факультет",
        default_training_center="2",
        discipline_code_prefix="BP",
        output_dir="docs/csv_templates/bp_draft",
        notes="Кафедра и ЛР — УЦ №2.",
        filename_hints=("безопасности производств", "кафедра 36", "(бп)"),
    ),
    "vd": DepartmentProfile(
        key="vd",
        department_code="ВД",
        department_title="Кафедра взрывного дела",
        faculty_code="ГФ",
        faculty_title="Горный факультет",
        default_training_center="2",
        discipline_code_prefix="VD",
        output_dir="docs/csv_templates/vd_draft",
        notes="Кафедра и ЛР — УЦ №2.",
        filename_hints=("взрывного дела", "кафедра 37", "(вд)"),
    ),
}


def norm(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def norm_key(value: str) -> str:
    return norm(value).lower()


def split_groups(raw: str) -> list[str]:
    if not raw:
        return []
    found = re.findall(r"[A-Za-zА-Яа-яЁё]{1,8}-\d+(?:-\d+)?", raw)
    if found:
        groups: list[str] = []
        for token in found:
            normalized = norm(token)
            if normalized and normalized not in groups:
                groups.append(normalized)
        return groups
    parts = re.split(r"[\n;,/]+", raw)
    groups = []
    for part in parts:
        token = norm(part)
        if token and token not in groups:
            groups.append(token)
    return groups


def parse_room(raw: str) -> str:
    text = norm(raw)
    if not text:
        return ""
    lowered = text.lower()
    if lowered in {"б/н", "нет", "—", "-", "нет аудитории", "без номера"}:
        return ""
    match = re.search(r"\d+(?:\.\d+)?", text.replace(",", "."))
    if not match:
        return ""
    number = match.group(0)
    if "." in number:
        number = number.split(".", 1)[0]
    return number.lstrip("0") or "0"


def parse_int(raw) -> str:
    if raw is None or raw == "":
        return ""
    if isinstance(raw, bool):
        return ""
    if isinstance(raw, (int, float)):
        return str(int(raw))
    text = norm(raw)
    if not text:
        return ""
    match = re.search(r"\d+", text)
    return match.group(0) if match else ""


def parse_duration(raw) -> str:
    value = parse_int(raw)
    if not value:
        return ""
    minutes = int(value)
    if minutes <= 0:
        return ""
    allowed = (45, 90, 135, 180, 270, 360)
    if minutes in allowed:
        return str(minutes)
    return str(min(allowed, key=lambda item: abs(item - minutes)))


def normalize_teacher_name(raw: str) -> str:
    teacher = norm(raw).split("\n")[0].strip()
    teacher = re.sub(r"[?\s]+$", "", teacher)
    if not teacher or teacher.lower() == "вакансия":
        return ""
    return teacher


def is_template_example(text: str) -> bool:
    """True for Excel template rows like «ПРИМЕР» or «… (ПРИМЕР)»."""
    value = norm_key(text)
    if not value:
        return False
    if value == "пример":
        return True
    return bool(re.search(r"\(\s*пример\s*\)", value))


def header_map(headers: list) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers, start=1):
        header_text = norm_key(str(header or ""))
        if not header_text:
            continue
        for field_name, aliases in COLUMN_ALIASES.items():
            if field_name in mapping:
                continue
            if any(alias in header_text for alias in aliases):
                mapping[field_name] = index
    return mapping


def detect_header_row(ws, *, max_row: int = 20) -> tuple[int, dict[str, int]] | None:
    for row_idx in range(1, max_row + 1):
        headers = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        columns = header_map(headers)
        if "lab_title" in columns and "discipline" in columns:
            return row_idx, columns
    return None


def should_skip_sheet(name: str) -> bool:
    lowered = norm_key(name)
    if any(lowered.startswith(prefix) for prefix in SKIP_SHEET_PREFIXES):
        return True
    return any(lowered.endswith(suffix.strip()) for suffix in SKIP_SHEET_SUFFIXES)


def sheet_year_rank(sheet_name: str) -> int:
    name = sheet_name.lower()
    for year, rank in (
        ("2026", 7),
        ("2025", 6),
        ("2024", 5),
        ("2023", 4),
        ("2022", 3),
        ("2021", 2),
    ):
        if year in name:
            return rank
    return 0


def resolve_profile(path: Path, explicit: str = "") -> DepartmentProfile:
    if explicit:
        if explicit not in DEPARTMENTS:
            raise SystemExit(f"Unknown department key: {explicit}. Choose: {', '.join(DEPARTMENTS)}")
        return DEPARTMENTS[explicit]
    lowered = path.name.lower()
    for profile in DEPARTMENTS.values():
        if any(hint in lowered for hint in profile.filename_hints):
            return profile
    raise SystemExit(f"Cannot detect department for {path.name}. Use --department.")


@dataclass
class ParsedRow:
    sheet: str
    discipline: str = ""
    groups: list[str] = field(default_factory=list)
    course: str = ""
    group_count: str = ""
    student_count: str = ""
    weeks: str = ""
    semester_dates: str = ""
    lab_hours: str = ""
    lab_title: str = ""
    equipment: str = ""
    capacity: str = ""
    duration_minutes: str = ""
    room: str = ""
    training_center: str = ""
    teacher: str = ""
    teacher_requires_presence: str = ""
    comments: str = ""


def parse_sheet(ws, sheet_name: str, profile: DepartmentProfile) -> list[ParsedRow]:
    detected = detect_header_row(ws)
    if not detected:
        return []
    header_row, columns = detected
    data_start = header_row + 1

    rows: list[ParsedRow] = []
    carry = {
        "discipline": "",
        "groups_raw": "",
        "course": "",
        "group_count": "",
        "student_count": "",
        "weeks": "",
        "semester_dates": "",
        "lab_hours": "",
        "teacher": "",
    }

    for row_idx in range(data_start, ws.max_row + 1):
        discipline_cell = ""
        if columns.get("discipline"):
            discipline_cell = norm(ws.cell(row_idx, columns["discipline"]).value)
            if discipline_cell and not is_template_example(discipline_cell):
                carry["discipline"] = discipline_cell

        effective_discipline = discipline_cell or carry["discipline"]

        if columns.get("groups_raw"):
            groups_raw = norm(ws.cell(row_idx, columns["groups_raw"]).value)
            if groups_raw:
                carry["groups_raw"] = groups_raw

        for key, col_name in (
            ("course", "course"),
            ("group_count", "group_count"),
            ("student_count", "student_count"),
            ("weeks", "weeks"),
            ("semester_dates", "semester_dates"),
            ("lab_hours", "lab_hours"),
        ):
            col = columns.get(col_name)
            if not col:
                continue
            raw = ws.cell(row_idx, col).value
            value = parse_int(raw) if key != "semester_dates" else norm(raw)
            if value:
                carry[key] = value

        if columns.get("teacher"):
            teacher = normalize_teacher_name(ws.cell(row_idx, columns["teacher"]).value)
            if teacher:
                carry["teacher"] = teacher

        lab_title = norm(ws.cell(row_idx, columns["lab_title"]).value)
        if not lab_title or is_template_example(lab_title):
            continue
        if is_template_example(effective_discipline):
            continue

        equipment = norm(ws.cell(row_idx, columns["equipment"]).value) if columns.get("equipment") else ""
        capacity = parse_int(ws.cell(row_idx, columns["capacity"]).value) if columns.get("capacity") else ""
        duration_minutes = (
            parse_duration(ws.cell(row_idx, columns["duration_minutes"]).value)
            if columns.get("duration_minutes")
            else ""
        )
        room = parse_room(ws.cell(row_idx, columns["room"]).value) if columns.get("room") else ""
        tc_from_file = (
            parse_int(ws.cell(row_idx, columns["training_center"]).value) if columns.get("training_center") else ""
        )
        training_center = tc_from_file or profile.default_training_center
        requires = (
            ws.cell(row_idx, columns["teacher_requires_presence"]).value
            if columns.get("teacher_requires_presence")
            else None
        )
        comments = norm(ws.cell(row_idx, columns["comments"]).value) if columns.get("comments") else ""

        groups = split_groups(carry["groups_raw"])
        if not effective_discipline and not groups:
            continue

        rows.append(
            ParsedRow(
                sheet=sheet_name,
                discipline=effective_discipline,
                groups=groups,
                course=carry["course"],
                group_count=carry["group_count"],
                student_count=carry["student_count"],
                weeks=carry["weeks"],
                semester_dates=carry["semester_dates"],
                lab_hours=carry["lab_hours"],
                lab_title=lab_title,
                equipment=equipment,
                capacity=capacity,
                duration_minutes=duration_minutes,
                room=room,
                training_center=training_center,
                teacher=carry["teacher"],
                teacher_requires_presence="да" if requires is True else ("нет" if requires is False else ""),
                comments=comments,
            )
        )
    return rows


def dedupe_catalog(rows: list[ParsedRow], profile: DepartmentProfile) -> list[dict[str, str]]:
    best: dict[tuple, dict[str, str]] = {}

    for row in rows:
        groups = row.groups or [""]
        for group in groups:
            key = (norm_key(row.discipline), group.upper(), norm_key(row.lab_title), row.room)
            payload = {
                "faculty_code": profile.faculty_code,
                "department_code": profile.department_code,
                "department_title": profile.department_title,
                "discipline_title": row.discipline,
                "group_name": group,
                "course": row.course,
                "student_count": row.student_count,
                "lab_title": row.lab_title,
                "equipment": row.equipment,
                "capacity": row.capacity,
                "duration_minutes": row.duration_minutes,
                "room_number": row.room,
                "training_center_number": row.training_center,
                "teacher_name": row.teacher,
                "teacher_requires_presence": row.teacher_requires_presence,
                "source_sheets": row.sheet,
                "comments": row.comments,
                "check_ok": "",
                "review_comment": "",
            }
            existing = best.get(key)
            row_rank = sheet_year_rank(row.sheet)
            existing_rank = sheet_year_rank(existing["source_sheets"].split("|")[-1]) if existing else -1
            if existing is None or row_rank >= existing_rank:
                if existing and row.sheet not in existing["source_sheets"]:
                    payload["source_sheets"] = existing["source_sheets"] + "|" + row.sheet
                best[key] = payload
            elif row.sheet not in existing["source_sheets"]:
                existing["source_sheets"] += "|" + row.sheet

    result = list(best.values())
    result.sort(key=lambda item: (item["discipline_title"], item["group_name"], item["lab_title"]))
    return result


def build_disciplines(catalog: list[dict[str, str]], profile: DepartmentProfile) -> list[dict[str, str]]:
    seen: dict[str, dict[str, str]] = {}
    prefix = profile.discipline_code_prefix
    for row in catalog:
        title = row["discipline_title"]
        if not title or title in seen:
            continue
        seen[title] = {
            "faculty_code": profile.faculty_code,
            "department_code": profile.department_code,
            "discipline_code_suggested": f"{prefix}-{len(seen)+1:03d}",
            "discipline_title": title,
            "training_center_number": row["training_center_number"] or profile.default_training_center,
            "source_sheets": row["source_sheets"],
            "check_ok": "",
            "review_comment": "",
        }
    return sorted(seen.values(), key=lambda item: item["discipline_title"])


def build_groups(catalog: list[dict[str, str]], profile: DepartmentProfile) -> list[dict[str, str]]:
    seen: dict[str, dict[str, str]] = {}
    for row in catalog:
        group = row["group_name"]
        if not group or group in seen:
            continue
        seen[group] = {
            "faculty_code": profile.faculty_code,
            "group_name": group,
            "student_count_suggested": row["student_count"],
            "course_suggested": row["course"],
            "source_sheets": row["source_sheets"],
            "check_ok": "",
            "review_comment": "",
        }
    return sorted(seen.values(), key=lambda item: item["group_name"])


def build_curriculum(catalog: list[dict[str, str]], disciplines: list[dict[str, str]]) -> list[dict[str, str]]:
    title_to_code = {row["discipline_title"]: row["discipline_code_suggested"] for row in disciplines}
    seen: set[tuple[str, str]] = set()
    rows: list[dict[str, str]] = []
    for row in catalog:
        group = row["group_name"]
        code = title_to_code.get(row["discipline_title"], "")
        if not group or not code:
            continue
        key = (group, code)
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "group_name": group,
                "discipline_code_suggested": code,
                "discipline_title": row["discipline_title"],
                "source_sheets": row["source_sheets"],
                "check_ok": "",
                "review_comment": "",
            }
        )
    return sorted(rows, key=lambda item: (item["group_name"], item["discipline_title"]))


def build_teachers(catalog: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: dict[str, dict] = {}
    for row in catalog:
        teacher = row["teacher_name"]
        if not teacher:
            continue
        if teacher not in seen:
            seen[teacher] = {
                "discipline_titles": {row["discipline_title"]},
                "source_sheets": {row["source_sheets"]},
            }
        else:
            seen[teacher]["discipline_titles"].add(row["discipline_title"])
            seen[teacher]["source_sheets"].add(row["source_sheets"])
    rows = [
        {
            "teacher_name": teacher,
            "role_suggested": "TEACHER",
            "discipline_titles": "|".join(sorted(payload["discipline_titles"])),
            "source_sheets": "|".join(sorted(payload["source_sheets"])),
            "check_ok": "",
            "review_comment": "",
        }
        for teacher, payload in seen.items()
    ]
    return sorted(rows, key=lambda item: item["teacher_name"])


def build_lab_works_unique(catalog: list[dict[str, str]]) -> list[dict[str, str]]:
    best: dict[tuple, dict[str, str]] = {}
    for row in catalog:
        key = (norm_key(row["discipline_title"]), norm_key(row["lab_title"]), row["room_number"])
        if key in best:
            continue
        best[key] = {
            "discipline_title": row["discipline_title"],
            "lab_title": row["lab_title"],
            "equipment": row["equipment"],
            "capacity": row["capacity"],
            "duration_minutes": row["duration_minutes"],
            "room_number": row["room_number"],
            "training_center_number": row["training_center_number"],
            "teacher_name": row["teacher_name"],
            "groups_sample": row["group_name"],
            "source_sheets": row["source_sheets"],
            "check_ok": "",
            "review_comment": "",
        }
    return sorted(best.values(), key=lambda item: (item["discipline_title"], item["lab_title"]))


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_outputs(
    *,
    input_path: Path,
    output_dir: Path,
    profile: DepartmentProfile,
    catalog: list[dict[str, str]],
    disciplines: list[dict[str, str]],
    groups: list[dict[str, str]],
    curriculum: list[dict[str, str]],
    teachers: list[dict[str, str]],
    lab_works_unique: list[dict[str, str]],
) -> None:
    write_csv(
        output_dir / "01_department.csv",
        [
            "faculty_code",
            "faculty_title",
            "department_code",
            "department_title",
            "default_training_center",
            "source_file",
            "check_ok",
            "review_comment",
        ],
        [
            {
                "faculty_code": profile.faculty_code,
                "faculty_title": profile.faculty_title,
                "department_code": profile.department_code,
                "department_title": profile.department_title,
                "default_training_center": profile.default_training_center,
                "source_file": input_path.name,
                "check_ok": "",
                "review_comment": "",
            }
        ],
    )
    write_csv(
        output_dir / "02_disciplines.csv",
        [
            "faculty_code",
            "department_code",
            "discipline_code_suggested",
            "discipline_title",
            "training_center_number",
            "source_sheets",
            "check_ok",
            "review_comment",
        ],
        disciplines,
    )
    write_csv(
        output_dir / "03_groups.csv",
        [
            "faculty_code",
            "group_name",
            "student_count_suggested",
            "course_suggested",
            "source_sheets",
            "check_ok",
            "review_comment",
        ],
        groups,
    )
    write_csv(
        output_dir / "04_lab_works.csv",
        [
            "faculty_code",
            "department_code",
            "department_title",
            "discipline_title",
            "group_name",
            "course",
            "student_count",
            "lab_title",
            "equipment",
            "capacity",
            "duration_minutes",
            "room_number",
            "training_center_number",
            "teacher_name",
            "teacher_requires_presence",
            "source_sheets",
            "comments",
            "check_ok",
            "review_comment",
        ],
        catalog,
    )
    write_csv(
        output_dir / "05_curriculum.csv",
        [
            "group_name",
            "discipline_code_suggested",
            "discipline_title",
            "source_sheets",
            "check_ok",
            "review_comment",
        ],
        curriculum,
    )
    write_csv(
        output_dir / "06_teachers.csv",
        [
            "teacher_name",
            "role_suggested",
            "discipline_titles",
            "source_sheets",
            "check_ok",
            "review_comment",
        ],
        teachers,
    )
    write_csv(
        output_dir / "07_lab_works_unique.csv",
        [
            "discipline_title",
            "lab_title",
            "equipment",
            "capacity",
            "duration_minutes",
            "room_number",
            "training_center_number",
            "teacher_name",
            "groups_sample",
            "source_sheets",
            "check_ok",
            "review_comment",
        ],
        lab_works_unique,
    )

    notes = profile.notes or "Черновые данные для наполнения и тестов."
    (output_dir / "README.md").write_text(
        "\n".join(
            [
                f"# {profile.department_title} — draft catalog",
                "",
                f"Источник: `{input_path.name}`",
                "",
                notes,
                "Даты семестров не используются — объединение по всем вкладкам учебных годов.",
                "Преподаватели «вакансия» отфильтровываются.",
                "Строки-шаблоны «ПРИМЕР» / «… (ПРИМЕР)» отфильтровываются.",
                "",
                f"УЦ по умолчанию: **№{profile.default_training_center}**.",
                "",
                "## Перегенерация",
                "",
                "```bash",
                f"python scripts/parse_department_workload_xlsx.py --department {profile.key} \"{input_path}\"",
                "```",
                "",
            ]
        ),
        encoding="utf-8",
    )


def parse_workbook_file(
    input_path: Path,
    profile: DepartmentProfile,
    *,
    output_dir: Path | None = None,
    verbose: bool = True,
) -> dict[str, int]:
    if not input_path.is_file():
        raise SystemExit(f"File not found: {input_path}")

    target_dir = output_dir or Path(profile.output_dir)
    workbook = openpyxl.load_workbook(input_path, data_only=True)
    all_rows: list[ParsedRow] = []
    sheet_stats: list[tuple[str, int]] = []

    for sheet_name in workbook.sheetnames:
        if should_skip_sheet(sheet_name):
            continue
        parsed = parse_sheet(workbook[sheet_name], sheet_name, profile)
        if parsed:
            sheet_stats.append((sheet_name, len(parsed)))
            all_rows.extend(parsed)

    catalog = dedupe_catalog(all_rows, profile)
    disciplines = build_disciplines(catalog, profile)
    groups = build_groups(catalog, profile)
    curriculum = build_curriculum(catalog, disciplines)
    teachers = build_teachers(catalog)
    lab_works_unique = build_lab_works_unique(catalog)

    write_outputs(
        input_path=input_path,
        output_dir=target_dir,
        profile=profile,
        catalog=catalog,
        disciplines=disciplines,
        groups=groups,
        curriculum=curriculum,
        teachers=teachers,
        lab_works_unique=lab_works_unique,
    )

    stats = {
        "sheets": len(sheet_stats),
        "disciplines": len(disciplines),
        "groups": len(groups),
        "lab_works": len(catalog),
        "lab_works_unique": len(lab_works_unique),
        "curriculum": len(curriculum),
        "teachers": len(teachers),
    }

    if verbose:
        print(f"\n=== {profile.department_title} ({input_path.name}) ===")
        for name, count in sheet_stats:
            print(f"  {name}: {count} LR rows")
        print(
            f"Catalog: disciplines={stats['disciplines']}, groups={stats['groups']}, "
            f"lab_works={stats['lab_works']}, lab_works_unique={stats['lab_works_unique']}, "
            f"curriculum={stats['curriculum']}, teachers={stats['teachers']}"
        )
        print(f"Output: {target_dir.resolve()}")

    return stats


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Parse department workload Excel into draft CSVs.")
    parser.add_argument("inputs", nargs="*", help="Excel file path(s)")
    parser.add_argument(
        "--department",
        choices=sorted(DEPARTMENTS),
        help="Department profile key (auto-detected from filename if omitted)",
    )
    parser.add_argument("--output", default="", help="Override output directory")
    parser.add_argument(
        "--all-defaults",
        action="store_true",
        help="Parse all known department files from Downloads (Windows paths)",
    )
    args = parser.parse_args(argv)

    if args.all_defaults:
        downloads = Path(r"d:/Users/Mayorov_IV/Downloads")
        jobs = [
            (downloads / "Кафедра 23 Металлургии.xlsx", "met"),
            (downloads / "Кафедра 19 Общей и технической физики (ОТФ).xlsx", "otf"),
            (
                downloads / "Кафедра 22 Химических технологий и переработки энергоносителей (ХТ и ПЭ).xlsx",
                "htpe",
            ),
            (downloads / "Кафедра 24 Общей и физической химии (ОФХ).xlsx", "ofh"),
            (downloads / "Кафедра 25 Обогащения полезных ископаемых (ОПИ).xlsx", "opi"),
            (downloads / "Кафедра 36 Безопасности производств (БП).xlsx", "bp"),
            (downloads / "Кафедра 37 Взрывного дела (ВД).xlsx", "vd"),
        ]
        for path, key in jobs:
            profile = DEPARTMENTS[key]
            out = Path(args.output) if args.output else None
            parse_workbook_file(path, profile, output_dir=out)
        return 0

    if not args.inputs:
        parser.error("Provide input file(s) or use --all-defaults")

    for input_arg in args.inputs:
        input_path = Path(input_arg)
        profile = resolve_profile(input_path, args.department or "")
        out = Path(args.output) if args.output else None
        parse_workbook_file(input_path, profile, output_dir=out)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
