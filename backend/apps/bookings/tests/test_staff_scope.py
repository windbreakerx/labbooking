import pytest

from django.test import Client
from django.utils import timezone
from rest_framework.test import APIClient

from apps.academics.models import Discipline, LabWork, Semester, StudentGroup
from apps.bookings.tests.conftest import create_lab_work, next_open_weekday_pair, seed_manual_booking
from apps.academics.querysets import (
    staff_disciplines_qs,
    staff_lab_works_qs,
    staff_managed_disciplines_qs,
    staff_students_qs,
)
from apps.bookings.models import Booking, BookingStatus, SupportTicket
from apps.bookings.services import BookingService, staff_booking_filter, staff_lab_filter
from apps.scheduling.models import (
    LabSession,
    LabSessionStatus,
    LabStand,
    Laboratory,
    Room,
    ScheduleEntry,
    TrainingCenter,
)
from apps.users.models import User, UserRole


@pytest.fixture
def semester(db):
    return Semester.objects.create(
        name="Staff Scope",
        start_date="2026-01-01",
        end_date="2026-12-31",
        is_active=True,
    )


@pytest.fixture
def own_tc():
    return TrainingCenter.objects.create(number=11, name="Своя лаборатория")


@pytest.fixture
def foreign_tc():
    return TrainingCenter.objects.create(number=12, name="Чужая лаборатория")


@pytest.fixture
def own_discipline(semester, own_tc):
    d = Discipline.objects.create(title="Своя дисциплина", semester=semester, is_published=True)
    d.training_centers.add(own_tc)
    return d


@pytest.fixture
def foreign_discipline(semester, foreign_tc):
    d = Discipline.objects.create(title="Чужая дисциплина", semester=semester, is_published=True)
    d.training_centers.add(foreign_tc)
    return d


@pytest.fixture
def own_lab_work(own_discipline, own_tc):
    lw = create_lab_work(
        own_discipline,
        number=1,
        title="Своя ЛР",
        duration_minutes=90,
        is_published=True,
    )
    lw.training_centers.add(own_tc)
    return lw


@pytest.fixture
def foreign_lab_work(foreign_discipline, foreign_tc):
    lw = create_lab_work(
        foreign_discipline,
        number=1,
        title="Чужая ЛР",
        duration_minutes=90,
        is_published=True,
    )
    lw.training_centers.add(foreign_tc)
    return lw


@pytest.fixture
def student_group(own_discipline):
    group = StudentGroup.objects.create(name="STAFF-SCOPE-24")
    group.disciplines.add(own_discipline)
    return group


@pytest.fixture
def student(db, student_group):
    user = User.objects.create_user(
        email="staff-scope-student@stud.spmi.ru",
        password="pass",
        first_name="Stu",
        last_name="Dent",
        role=UserRole.STUDENT,
    )
    user.profile.student_group = student_group
    user.profile.save(update_fields=["student_group"])
    return user


@pytest.fixture
def staff_with_lab(db, own_tc):
    user = User.objects.create_user(
        email="staff-scope@spmi.ru",
        password="pass",
        first_name="Own",
        last_name="Staff",
        role=UserRole.LAB_ADMIN,
        is_staff=True,
    )
    user.profile.training_center = own_tc
    user.profile.save(update_fields=["training_center"])
    return user


@pytest.fixture
def staff_no_lab(db):
    return User.objects.create_user(
        email="staff-nolab@spmi.ru",
        password="pass",
        first_name="No",
        last_name="Lab",
        role=UserRole.LAB_ADMIN,
        is_staff=True,
    )


@pytest.fixture
def lab_head(db, own_tc):
    user = User.objects.create_user(
        email="lab-head@spmi.ru",
        password="pass",
        first_name="Head",
        last_name="Lab",
        role=UserRole.LAB_HEAD,
        is_staff=True,
    )
    user.profile.training_center = own_tc
    user.profile.save(update_fields=["training_center"])
    return user


@pytest.fixture
def foreign_staff(db, foreign_tc):
    user = User.objects.create_user(
        email="foreign-staff@spmi.ru",
        password="pass",
        first_name="Foreign",
        last_name="Staff",
        role=UserRole.LAB_ADMIN,
        is_staff=True,
    )
    user.profile.training_center = foreign_tc
    user.profile.save(update_fields=["training_center"])
    return user


@pytest.fixture
def own_teacher(db, own_tc):
    user = User.objects.create_user(
        email="own-teacher@spmi.ru",
        password="pass",
        first_name="Own",
        last_name="Teacher",
        role=UserRole.TEACHER,
    )
    user.profile.training_center = own_tc
    user.profile.save(update_fields=["training_center"])
    return user


@pytest.fixture
def sys_admin(db):
    return User.objects.create_user(
        email="sysadmin@spmi.ru",
        password="pass",
        first_name="Sys",
        last_name="Admin",
        role=UserRole.SYS_ADMIN,
        is_staff=True,
        is_superuser=True,
    )


@pytest.fixture
def own_room(own_tc):
    return Room.objects.create(training_center=own_tc, number="110", capacity=5)


@pytest.fixture
def foreign_room(foreign_tc):
    return Room.objects.create(training_center=foreign_tc, number="220", capacity=5)


@pytest.fixture
def own_session(own_lab_work, own_room, semester):
    starts = next_open_weekday_pair(days_ahead=7)
    return LabSession.objects.create(
        lab_work=own_lab_work,
        room=own_room,
        semester=semester,
        starts_at=starts,
        ends_at=starts + timezone.timedelta(minutes=90),
        capacity=5,
        status=LabSessionStatus.OPEN,
    )


@pytest.fixture
def foreign_session(foreign_lab_work, foreign_room, semester):
    starts = next_open_weekday_pair(days_ahead=8)
    return LabSession.objects.create(
        lab_work=foreign_lab_work,
        room=foreign_room,
        semester=semester,
        starts_at=starts,
        ends_at=starts + timezone.timedelta(minutes=90),
        capacity=5,
        status=LabSessionStatus.OPEN,
    )


@pytest.fixture
def own_booking(staff_with_lab, student, own_session):
    return BookingService(actor=staff_with_lab).create_booking(
        student,
        own_session.pk,
        manual=True,
    )


@pytest.fixture
def foreign_booking(staff_with_lab, student, foreign_session):
    return seed_manual_booking(actor=staff_with_lab, student=student, session=foreign_session)


@pytest.fixture
def own_ticket(student, own_tc):
    return SupportTicket.objects.create(
        student=student,
        subject="Своё обращение",
        body="Текст",
        training_center=own_tc,
    )


@pytest.fixture
def foreign_ticket(student, foreign_tc):
    return SupportTicket.objects.create(
        student=student,
        subject="Чужое обращение",
        body="Текст",
        training_center=foreign_tc,
    )


@pytest.fixture
def own_stand(own_room, own_tc):
    return LabStand.objects.create(
        name="Свой стенд",
        inventory_number="OWN-1",
        training_center=own_tc,
        room=own_room,
    )


@pytest.fixture
def foreign_stand(foreign_room, foreign_tc):
    return LabStand.objects.create(
        name="Чужой стенд",
        inventory_number="FOR-1",
        training_center=foreign_tc,
        room=foreign_room,
    )


@pytest.fixture
def own_schedule(own_lab_work, own_room, semester):
    return ScheduleEntry.objects.create(
        lab_work=own_lab_work,
        room=own_room,
        semester=semester,
        weekday=0,
        start_time="10:35",
    )


@pytest.fixture
def foreign_schedule(foreign_lab_work, foreign_room, semester):
    return ScheduleEntry.objects.create(
        lab_work=foreign_lab_work,
        room=foreign_room,
        semester=semester,
        weekday=1,
        start_time="12:20",
    )


@pytest.mark.django_db
class TestStaffScopeQuerysets:
    def test_staff_without_lab_sees_nothing(self, staff_no_lab, own_discipline, foreign_discipline):
        assert staff_disciplines_qs(staff_no_lab).count() == 0
        assert staff_managed_disciplines_qs(staff_no_lab).count() == 0
        assert staff_lab_filter(Discipline.objects.all(), staff_no_lab, training_center_lookup="training_centers").count() == 0

    def test_staff_lab_filter_scoped(self, staff_with_lab, own_booking, foreign_booking):
        from apps.bookings.models import Booking

        ids = set(staff_lab_filter(Booking.objects.all(), staff_with_lab).values_list("pk", flat=True))
        assert own_booking.pk in ids
        assert foreign_booking.pk not in ids

    def test_lab_head_has_staff_scope(self, lab_head, own_discipline, foreign_discipline):
        ids = set(staff_disciplines_qs(lab_head).values_list("pk", flat=True))
        assert own_discipline.pk in ids
        assert foreign_discipline.pk not in ids

    def test_sys_admin_sees_all_disciplines(self, sys_admin, own_discipline, foreign_discipline):
        ids = set(staff_disciplines_qs(sys_admin).values_list("pk", flat=True))
        assert own_discipline.pk in ids
        assert foreign_discipline.pk in ids


@pytest.mark.django_db
class TestStaffScopeWeb:
    def test_staff_disciplines_hides_foreign(self, staff_with_lab, own_discipline, foreign_discipline):
        client = Client()
        client.force_login(staff_with_lab)
        response = client.get("/staff/disciplines/")
        assert response.status_code == 200
        assert own_discipline.title.encode() in response.content
        assert foreign_discipline.title.encode() not in response.content

    def test_staff_lab_works_hides_foreign(self, staff_with_lab, own_lab_work, foreign_lab_work):
        client = Client()
        client.force_login(staff_with_lab)
        response = client.get("/staff/lab-works/")
        assert response.status_code == 200
        assert own_lab_work.title.encode() in response.content
        assert foreign_lab_work.title.encode() not in response.content

    def test_staff_bookings_hides_foreign(self, staff_with_lab, own_booking, foreign_booking, foreign_room):
        client = Client()
        client.force_login(staff_with_lab)
        response = client.get("/staff/bookings/")
        assert response.status_code == 200
        assert foreign_room.number.encode() not in response.content

    def test_staff_support_hides_foreign(self, staff_with_lab, own_ticket, foreign_ticket):
        client = Client()
        client.force_login(staff_with_lab)
        response = client.get("/staff/support/")
        assert response.status_code == 200
        assert own_ticket.subject.encode() in response.content
        assert foreign_ticket.subject.encode() not in response.content

    def test_staff_support_reply_foreign_404(self, staff_with_lab, foreign_ticket):
        client = Client()
        client.force_login(staff_with_lab)
        response = client.post(
            f"/staff/support/{foreign_ticket.pk}/reply/",
            {"body": "Ответ"},
        )
        assert response.status_code == 404

    def test_staff_stands_hides_foreign(self, staff_with_lab, own_stand, foreign_stand):
        client = Client()
        client.force_login(staff_with_lab)
        response = client.get("/staff/stands/")
        assert response.status_code == 200
        assert own_stand.name.encode() in response.content
        assert foreign_stand.name.encode() not in response.content

    def test_staff_schedule_stub(self, staff_with_lab, own_schedule, foreign_schedule):
        client = Client()
        client.force_login(staff_with_lab)
        response = client.get("/staff/schedule/")
        assert response.status_code == 200
        content = response.content.decode()
        assert "На доработке" in content
        assert own_schedule.lab_work.title not in content
        assert foreign_schedule.lab_work.title not in content

    def test_staff_people_hides_foreign_lab_people(
        self,
        staff_with_lab,
        own_teacher,
        foreign_staff,
    ):
        client = Client()
        client.force_login(staff_with_lab)
        response = client.get("/staff/people/")
        assert response.status_code == 200
        assert own_teacher.email.encode() in response.content
        assert foreign_staff.email.encode() not in response.content

    def test_staff_students_shows_group_and_booking(
        self,
        staff_with_lab,
        student,
        student_group,
        own_booking,
    ):
        client = Client()
        client.force_login(staff_with_lab)
        response = client.get("/staff/students/")
        assert response.status_code == 200
        assert student.email.encode() in response.content
        assert student_group.name.encode() in response.content
        assert own_booking.lab_work.title.encode() in response.content

    def test_staff_students_hides_foreign_group(
        self,
        staff_with_lab,
        foreign_discipline,
        semester,
    ):
        foreign_group = StudentGroup.objects.create(name="FOREIGN-GROUP-24")
        foreign_group.disciplines.add(foreign_discipline)
        foreign_student = User.objects.create_user(
            email="foreign-student@stud.spmi.ru",
            password="pass",
            first_name="Foreign",
            last_name="Student",
            role=UserRole.STUDENT,
        )
        foreign_student.profile.student_group = foreign_group
        foreign_student.profile.save(update_fields=["student_group"])

        client = Client()
        client.force_login(staff_with_lab)
        response = client.get("/staff/students/")
        assert response.status_code == 200
        assert foreign_student.email.encode() not in response.content

    def test_staff_students_scoped_queryset(
        self,
        staff_with_lab,
        student,
        foreign_discipline,
        semester,
    ):
        ids = set(staff_students_qs(staff_with_lab).values_list("pk", flat=True))
        assert student.pk in ids

        foreign_group = StudentGroup.objects.create(name="FOREIGN-QS-24")
        foreign_group.disciplines.add(foreign_discipline)
        foreign_student = User.objects.create_user(
            email="foreign-qs@stud.spmi.ru",
            password="pass",
            first_name="Foreign",
            last_name="QS",
            role=UserRole.STUDENT,
        )
        foreign_student.profile.student_group = foreign_group
        foreign_student.profile.save(update_fields=["student_group"])
        ids = set(staff_students_qs(staff_with_lab).values_list("pk", flat=True))
        assert foreign_student.pk not in ids

    def test_staff_students_filter_by_group(self, staff_with_lab, student, student_group):
        client = Client()
        client.force_login(staff_with_lab)
        response = client.get("/staff/students/", {"group": student_group.name})
        assert response.status_code == 200
        assert student.email.encode() in response.content

    def test_staff_students_sort_by_email(self, staff_with_lab, student, student_group):
        other = User.objects.create_user(
            email="aaa-student@stud.spmi.ru",
            password="pass",
            first_name="А",
            last_name="Яя",
            role=UserRole.STUDENT,
        )
        other.profile.student_group = student_group
        other.profile.save(update_fields=["student_group"])

        client = Client()
        client.force_login(staff_with_lab)
        response = client.get("/staff/students/", {"sort": "email", "dir": "asc"})
        assert response.status_code == 200
        content = response.content.decode()
        assert content.index("aaa-student@stud.spmi.ru") < content.index(student.email)

    def test_staff_no_lab_empty_pages(self, staff_no_lab, own_discipline, foreign_discipline, own_ticket):
        client = Client()
        client.force_login(staff_no_lab)
        for url in (
            "/staff/disciplines/",
            "/staff/lab-works/",
            "/staff/bookings/",
            "/staff/support/",
            "/staff/stands/",
            "/staff/schedule/",
            "/staff/people/",
            "/staff/students/",
        ):
            response = client.get(url)
            assert response.status_code == 200
        assert own_discipline.title.encode() not in client.get("/staff/disciplines/").content
        assert own_ticket.subject.encode() not in client.get("/staff/support/").content

    def test_sys_admin_people_sees_all_labs(self, sys_admin, own_teacher, foreign_staff):
        client = Client()
        client.force_login(sys_admin)
        response = client.get("/staff/people/")
        assert response.status_code == 200
        assert own_teacher.email.encode() in response.content
        assert foreign_staff.email.encode() in response.content

    def test_staff_status_foreign_booking_blocked(
        self,
        staff_with_lab,
        foreign_booking,
    ):
        client = Client()
        client.force_login(staff_with_lab)
        response = client.post(
            f"/staff/bookings/{foreign_booking.pk}/status/",
            {"status": BookingStatus.VISITED},
        )
        assert response.status_code == 302
        foreign_booking.refresh_from_db()
        assert foreign_booking.current_status != BookingStatus.VISITED


@pytest.mark.django_db
class TestTeacherBookingReadOnly:
    def test_teacher_get_bookings_web(self, own_teacher, own_booking):
        client = Client()
        client.force_login(own_teacher)
        response = client.get("/staff/bookings/")
        assert response.status_code == 200
        assert own_booking.student.email.encode() in response.content

    def test_teacher_post_status_forbidden(self, own_teacher, own_booking):
        client = Client()
        client.force_login(own_teacher)
        response = client.post(
            f"/staff/bookings/{own_booking.pk}/status/",
            {"status": BookingStatus.VISITED},
        )
        assert response.status_code == 403
        own_booking.refresh_from_db()
        assert own_booking.current_status == BookingStatus.BOOKED

    def test_teacher_get_bookings_api(self, own_teacher, own_booking):
        client = APIClient()
        client.force_authenticate(user=own_teacher)
        response = client.get("/api/v1/admin/bookings/")
        assert response.status_code == 200
        ids = {item["id"] for item in response.json()["results"]}
        assert own_booking.pk in ids

    def test_teacher_patch_status_api_forbidden(self, own_teacher, own_booking):
        client = APIClient()
        client.force_authenticate(user=own_teacher)
        response = client.patch(
            f"/api/v1/admin/bookings/{own_booking.pk}/status/",
            {"status": BookingStatus.VISITED},
            format="json",
        )
        assert response.status_code == 403
        own_booking.refresh_from_db()
        assert own_booking.current_status == BookingStatus.BOOKED

    def test_teacher_manual_booking_api_forbidden(self, own_teacher, student, own_session):
        client = APIClient()
        client.force_authenticate(user=own_teacher)
        response = client.post(
            "/api/v1/admin/bookings/manual/",
            {"student_id": student.pk, "lab_session_id": own_session.pk},
            format="json",
        )
        assert response.status_code == 403

    def test_teacher_report_still_allowed(self, own_teacher, own_booking):
        client = APIClient()
        client.force_authenticate(user=own_teacher)
        response = client.get("/api/v1/admin/reports/bookings/")
        assert response.status_code == 200

    def test_lab_admin_post_status_allowed(self, staff_with_lab, own_booking):
        client = Client()
        client.force_login(staff_with_lab)
        response = client.post(
            f"/staff/bookings/{own_booking.pk}/status/",
            {"status": BookingStatus.VISITED},
        )
        assert response.status_code == 302
        own_booking.refresh_from_db()
        assert own_booking.current_status == BookingStatus.VISITED

    def test_lab_head_post_status_allowed(self, lab_head, own_booking):
        client = Client()
        client.force_login(lab_head)
        response = client.post(
            f"/staff/bookings/{own_booking.pk}/status/",
            {"status": BookingStatus.VISITED},
        )
        assert response.status_code == 302
        own_booking.refresh_from_db()
        assert own_booking.current_status == BookingStatus.VISITED


@pytest.mark.django_db
class TestStaffScopeApi:
    def test_admin_bookings_scoped(self, staff_with_lab, own_booking, foreign_booking):
        client = APIClient()
        client.force_authenticate(user=staff_with_lab)
        response = client.get("/api/v1/admin/bookings/")
        assert response.status_code == 200
        ids = {item["id"] for item in response.json()["results"]}
        assert own_booking.pk in ids
        assert foreign_booking.pk not in ids

    def test_disciplines_api_scoped(self, staff_with_lab, own_discipline, foreign_discipline):
        client = APIClient()
        client.force_authenticate(user=staff_with_lab)
        response = client.get("/api/v1/disciplines/")
        assert response.status_code == 200
        ids = {item["id"] for item in response.json()["results"]}
        assert own_discipline.pk in ids
        assert foreign_discipline.pk not in ids

    def test_foreign_discipline_lab_works_api_404(self, staff_with_lab, foreign_discipline):
        client = APIClient()
        client.force_authenticate(user=staff_with_lab)
        response = client.get(f"/api/v1/disciplines/{foreign_discipline.pk}/lab-works/")
        assert response.status_code == 404

    def test_support_tickets_scoped(self, staff_with_lab, own_ticket, foreign_ticket):
        client = APIClient()
        client.force_authenticate(user=staff_with_lab)
        response = client.get("/api/v1/support/tickets/")
        assert response.status_code == 200
        ids = {item["id"] for item in response.json()["results"]}
        assert own_ticket.pk in ids
        assert foreign_ticket.pk not in ids

    def test_support_message_foreign_ticket_denied(self, staff_with_lab, foreign_ticket):
        client = APIClient()
        client.force_authenticate(user=staff_with_lab)
        response = client.post(
            f"/api/v1/support/tickets/{foreign_ticket.pk}/messages/",
            {"body": "Ответ"},
            format="json",
        )
        assert response.status_code == 403

    def test_admin_sessions_scoped(self, staff_with_lab, own_session, foreign_session):
        client = APIClient()
        client.force_authenticate(user=staff_with_lab)
        response = client.get("/api/v1/admin/sessions/")
        assert response.status_code == 200
        ids = {item["id"] for item in response.json()["results"]}
        assert own_session.pk in ids
        assert foreign_session.pk not in ids

    def test_status_update_foreign_booking_denied(self, staff_with_lab, foreign_booking):
        client = APIClient()
        client.force_authenticate(user=staff_with_lab)
        response = client.patch(
            f"/api/v1/admin/bookings/{foreign_booking.pk}/status/",
            {"status": BookingStatus.VISITED},
            format="json",
        )
        assert response.status_code == 403

    def test_manual_booking_foreign_session_denied(self, staff_with_lab, student, foreign_session):
        client = APIClient()
        client.force_authenticate(user=staff_with_lab)
        response = client.post(
            "/api/v1/admin/bookings/manual/",
            {"student_id": student.pk, "lab_session_id": foreign_session.pk},
            format="json",
        )
        assert response.status_code == 403

    def test_staff_no_lab_empty_api_lists(self, staff_no_lab, own_booking, own_discipline):
        client = APIClient()
        client.force_authenticate(user=staff_no_lab)
        bookings = client.get("/api/v1/admin/bookings/")
        disciplines = client.get("/api/v1/disciplines/")
        sessions = client.get("/api/v1/admin/sessions/")
        assert bookings.status_code == 200
        assert disciplines.status_code == 200
        assert sessions.status_code == 200
        assert bookings.json()["results"] == []
        assert disciplines.json()["results"] == []
        assert sessions.json()["results"] == []

    def test_report_excludes_foreign_booking(self, staff_with_lab, own_booking, foreign_booking):
        client = APIClient()
        client.force_authenticate(user=staff_with_lab)
        response = client.get("/api/v1/admin/reports/bookings/")
        assert response.status_code == 200
        from openpyxl import load_workbook
        from io import BytesIO

        wb = load_workbook(BytesIO(response.content))
        rows = list(wb.active.iter_rows(values_only=True))
        assert len(rows) == 2  # заголовок + одна запись своей лаборатории
        assert foreign_booking.discipline.title not in {row[3] for row in rows[1:]}

    def test_lab_works_scoped_in_queryset(self, staff_with_lab, own_lab_work, foreign_lab_work):
        ids = set(staff_lab_works_qs(staff_with_lab).values_list("pk", flat=True))
        assert own_lab_work.pk in ids
        assert foreign_lab_work.pk not in ids


@pytest.fixture
def shared_tc(db):
    return TrainingCenter.objects.create(number=50, name="Shared UC")


@pytest.fixture
def lab_a(shared_tc):
    return Laboratory.objects.create(training_center=shared_tc, name="Lab A")


@pytest.fixture
def lab_b(shared_tc):
    return Laboratory.objects.create(training_center=shared_tc, name="Lab B")


@pytest.fixture
def staff_lab_a(db, shared_tc, lab_a):
    user = User.objects.create_user(
        email="staff-lab-a@spmi.ru",
        password="pass",
        first_name="Staff",
        last_name="LabA",
        role=UserRole.LAB_ADMIN,
        is_staff=True,
    )
    user.profile.training_center = shared_tc
    user.profile.laboratory = lab_a
    user.profile.save(update_fields=["training_center", "laboratory"])
    return user


@pytest.fixture
def lab_a_discipline(semester, shared_tc, lab_a):
    discipline = Discipline.objects.create(
        title="Lab A discipline",
        semester=semester,
        is_published=True,
    )
    discipline.training_centers.add(shared_tc)
    discipline.laboratories.add(lab_a)
    return discipline


@pytest.fixture
def lab_b_discipline(semester, shared_tc, lab_b):
    discipline = Discipline.objects.create(
        title="Lab B discipline",
        semester=semester,
        is_published=True,
    )
    discipline.training_centers.add(shared_tc)
    discipline.laboratories.add(lab_b)
    return discipline


@pytest.fixture
def lab_a_lab_work(lab_a_discipline, shared_tc, lab_a):
    lab_work = create_lab_work(
        lab_a_discipline,
        number=1,
        title="Lab A LR",
        duration_minutes=90,
        is_published=True,
    )
    lab_work.training_centers.add(shared_tc)
    lab_work.laboratories.add(lab_a)
    return lab_work


@pytest.fixture
def lab_b_lab_work(lab_b_discipline, shared_tc, lab_b):
    lab_work = create_lab_work(
        lab_b_discipline,
        number=1,
        title="Lab B LR",
        duration_minutes=90,
        is_published=True,
    )
    lab_work.training_centers.add(shared_tc)
    lab_work.laboratories.add(lab_b)
    return lab_work


@pytest.fixture
def room_a(shared_tc, lab_a):
    return Room.objects.create(
        training_center=shared_tc,
        laboratory=lab_a,
        number="A-101",
        capacity=5,
    )


@pytest.fixture
def room_b(shared_tc, lab_b):
    return Room.objects.create(
        training_center=shared_tc,
        laboratory=lab_b,
        number="B-201",
        capacity=5,
    )


@pytest.fixture
def session_a(lab_a_lab_work, room_a, semester):
    starts = next_open_weekday_pair(days_ahead=7)
    return LabSession.objects.create(
        lab_work=lab_a_lab_work,
        room=room_a,
        semester=semester,
        starts_at=starts,
        ends_at=starts + timezone.timedelta(minutes=90),
        capacity=5,
        status=LabSessionStatus.OPEN,
    )


@pytest.fixture
def session_b(lab_b_lab_work, room_b, semester):
    starts = next_open_weekday_pair(days_ahead=8)
    return LabSession.objects.create(
        lab_work=lab_b_lab_work,
        room=room_b,
        semester=semester,
        starts_at=starts,
        ends_at=starts + timezone.timedelta(minutes=90),
        capacity=5,
        status=LabSessionStatus.OPEN,
    )


@pytest.fixture
def booking_a(staff_lab_a, student, session_a):
    return seed_manual_booking(actor=staff_lab_a, student=student, session=session_a)


@pytest.fixture
def booking_b(staff_lab_a, student, session_b):
    return seed_manual_booking(actor=staff_lab_a, student=student, session=session_b)


@pytest.mark.django_db
class TestStaffLaboratoryIsolation:
    def test_staff_lab_filter_uses_laboratory_when_profile_set(
        self,
        staff_lab_a,
        booking_a,
        booking_b,
    ):
        ids = set(staff_lab_filter(Booking.objects.all(), staff_lab_a).values_list("pk", flat=True))
        assert booking_a.pk in ids
        assert booking_b.pk not in ids

    def test_staff_booking_filter_keeps_lab_work_when_room_laboratory_differs(
        self,
        staff_lab_a,
        student,
        lab_a_lab_work,
        lab_a_discipline,
        room_b,
        semester,
    ):
        starts = next_open_weekday_pair(days_ahead=10)
        session = LabSession.objects.create(
            lab_work=lab_a_lab_work,
            room=room_b,
            semester=semester,
            starts_at=starts,
            ends_at=starts + timezone.timedelta(minutes=90),
            capacity=5,
            status=LabSessionStatus.OPEN,
        )
        booking = seed_manual_booking(actor=staff_lab_a, student=student, session=session)
        ids = set(staff_booking_filter(Booking.objects.all(), staff_lab_a).values_list("pk", flat=True))
        assert booking.pk in ids
        strict_ids = set(staff_lab_filter(Booking.objects.all(), staff_lab_a).values_list("pk", flat=True))
        assert booking.pk not in strict_ids

    def test_staff_students_page_with_lab_work_room_mismatch(
        self,
        staff_lab_a,
        student,
        lab_a_lab_work,
        room_b,
        semester,
    ):
        starts = next_open_weekday_pair(days_ahead=11)
        session = LabSession.objects.create(
            lab_work=lab_a_lab_work,
            room=room_b,
            semester=semester,
            starts_at=starts,
            ends_at=starts + timezone.timedelta(minutes=90),
            capacity=5,
            status=LabSessionStatus.OPEN,
        )
        seed_manual_booking(actor=staff_lab_a, student=student, session=session)
        client = Client()
        client.force_login(staff_lab_a)
        response = client.get("/staff/students/")
        assert response.status_code == 200
        assert student.email.encode() in response.content

    def test_staff_bookings_page_orders_after_lab_work_scope_fallback(
        self,
        staff_lab_a,
        student,
        lab_a_lab_work,
        room_b,
        semester,
    ):
        starts = next_open_weekday_pair(days_ahead=12)
        session = LabSession.objects.create(
            lab_work=lab_a_lab_work,
            room=room_b,
            semester=semester,
            starts_at=starts,
            ends_at=starts + timezone.timedelta(minutes=90),
            capacity=5,
            status=LabSessionStatus.OPEN,
        )
        seed_manual_booking(actor=staff_lab_a, student=student, session=session)
        client = Client()
        client.force_login(staff_lab_a)
        response = client.get("/staff/bookings/", {"sort": "student", "dir": "asc"})
        assert response.status_code == 200
        assert student.last_name.encode() in response.content

    def test_staff_bookings_web_hides_sibling_laboratory(
        self,
        staff_lab_a,
        booking_a,
        booking_b,
        room_b,
    ):
        client = Client()
        client.force_login(staff_lab_a)
        response = client.get("/staff/bookings/")
        assert response.status_code == 200
        assert booking_a.room.number.encode() in response.content
        assert room_b.number.encode() not in response.content

    def test_manual_booking_foreign_laboratory_session_denied(
        self,
        staff_lab_a,
        student,
        session_b,
    ):
        client = APIClient()
        client.force_authenticate(user=staff_lab_a)
        response = client.post(
            "/api/v1/admin/bookings/manual/",
            {"student_id": student.pk, "lab_session_id": session_b.pk},
            format="json",
        )
        assert response.status_code == 403

    def test_admin_report_excludes_sibling_laboratory_booking(
        self,
        staff_lab_a,
        booking_a,
        booking_b,
    ):
        client = APIClient()
        client.force_authenticate(user=staff_lab_a)
        response = client.get("/api/v1/admin/reports/bookings/")
        assert response.status_code == 200
        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(response.content))
        rows = list(wb.active.iter_rows(values_only=True))
        assert len(rows) == 2
        assert booking_b.discipline.title not in {row[3] for row in rows[1:]}

    def test_status_update_foreign_laboratory_booking_denied(
        self,
        staff_lab_a,
        booking_b,
    ):
        client = APIClient()
        client.force_authenticate(user=staff_lab_a)
        response = client.patch(
            f"/api/v1/admin/bookings/{booking_b.pk}/status/",
            {"status": BookingStatus.VISITED},
            format="json",
        )
        assert response.status_code == 403
        booking_b.refresh_from_db()
        assert booking_b.current_status != BookingStatus.VISITED

    def test_staff_people_hides_sibling_laboratory(
        self,
        staff_lab_a,
        lab_a,
        lab_b,
    ):
        teacher_a = User.objects.create_user(
            email="teacher-lab-a@spmi.ru",
            password="pass",
            first_name="Teacher",
            last_name="LabA",
            role=UserRole.TEACHER,
        )
        teacher_a.profile.training_center = lab_a.training_center
        teacher_a.profile.laboratory = lab_a
        teacher_a.profile.save(update_fields=["training_center", "laboratory"])

        teacher_b = User.objects.create_user(
            email="teacher-lab-b@spmi.ru",
            password="pass",
            first_name="Teacher",
            last_name="LabB",
            role=UserRole.TEACHER,
        )
        teacher_b.profile.training_center = lab_b.training_center
        teacher_b.profile.laboratory = lab_b
        teacher_b.profile.save(update_fields=["training_center", "laboratory"])

        client = Client()
        client.force_login(staff_lab_a)
        response = client.get("/staff/people/")
        assert response.status_code == 200
        assert teacher_a.email.encode() in response.content
        assert teacher_b.email.encode() not in response.content


@pytest.fixture
def lab_head_with_laboratory(db, shared_tc, lab_a):
    user = User.objects.create_user(
        email="lab-head-lab@spmi.ru",
        password="pass",
        first_name="Head",
        last_name="LabA",
        role=UserRole.LAB_HEAD,
        is_staff=True,
    )
    user.profile.training_center = shared_tc
    user.profile.laboratory = lab_a
    user.profile.save(update_fields=["training_center", "laboratory"])
    return user


@pytest.fixture
def room_a_unassigned(shared_tc):
    return Room.objects.create(
        training_center=shared_tc,
        number="A-legacy",
        capacity=5,
    )


@pytest.fixture
def session_a_unassigned(lab_a_lab_work, room_a_unassigned, semester):
    starts = next_open_weekday_pair(days_ahead=9)
    return LabSession.objects.create(
        lab_work=lab_a_lab_work,
        room=room_a_unassigned,
        semester=semester,
        starts_at=starts,
        ends_at=starts + timezone.timedelta(minutes=90),
        capacity=5,
        status=LabSessionStatus.OPEN,
    )


@pytest.fixture
def own_ticket_lab_a(student, shared_tc):
    return SupportTicket.objects.create(
        student=student,
        subject="Обращение в УЦ",
        body="Текст",
        training_center=shared_tc,
    )


@pytest.mark.django_db
class TestStaffLaboratoryScopeRegression:
    def test_lab_head_support_page_with_laboratory_profile(
        self,
        lab_head_with_laboratory,
        own_ticket_lab_a,
    ):
        client = Client()
        client.force_login(lab_head_with_laboratory)
        response = client.get("/staff/support/")
        assert response.status_code == 200
        assert own_ticket_lab_a.subject.encode() in response.content

    def test_bookings_visible_when_room_laboratory_unassigned(
        self,
        staff_lab_a,
        student,
        session_a_unassigned,
    ):
        booking = seed_manual_booking(
            actor=staff_lab_a,
            student=student,
            session=session_a_unassigned,
        )
        client = Client()
        client.force_login(staff_lab_a)
        response = client.get("/staff/bookings/")
        assert response.status_code == 200
        assert session_a_unassigned.room.number.encode() in response.content
        ids = set(staff_lab_filter(Booking.objects.all(), staff_lab_a).values_list("pk", flat=True))
        assert booking.pk in ids

    def test_staff_rooms_page_with_laboratory_profile(self, staff_lab_a, room_a, room_b):
        client = Client()
        client.force_login(staff_lab_a)
        response = client.get("/staff/rooms/")
        assert response.status_code == 200
        assert room_a.number.encode() in response.content
        assert room_b.number.encode() not in response.content
