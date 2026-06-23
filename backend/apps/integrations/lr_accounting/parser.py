"""Парсер Excel-журналов учёта лабораторных работ (ЛР_учет *.xlsx)."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.chartsheet import Chartsheet

SKIP_SHEETS = {"Свод", "Summary", "Диаграмма1", "Эффект", "Перечень лр", "Дежурство"}

FILE_ROOM_BY_PATTERN = (
    (re.compile(r"1123"), "1123"),
    (re.compile(r"2114-16"), "2114"),
    (re.compile(r"2116-18"), "2118"),
    (re.compile(r"2115"), "2115"),
    (re.compile(r"2117"), "2117"),
)

LAB_TITLE_PREFIXES = (
    "Лаб. раб.",
    "Изучение",
    "Исследование",
    "Определение",
    "Лаборатор",
    "Подготовка",
    "Настройка",
    "Перевод",
    "Запуск",
    "Динамометрирование",
    "Волнометрирование",
    "Капиллярный",
    "Магнитопорошковый",
    "Анализ",
    "Выявление",
    "Вихретоковая",
    "Вибрационная",
    "Ультразвуковой",
    "Измерение",
    "Лазерная",
)


@dataclass
class ParsedLabWork:
    discipline: str
    title: str
    duration_minutes: int | None = None
    catalog_number: int | None = None
    stand_name: str = ""


@dataclass
class ParsedStudent:
    number: int
    last_name: str
    first_name: str


@dataclass
class ParsedGroupSheet:
    name: str
    lab_works: list[ParsedLabWork] = field(default_factory=list)
    students: list[ParsedStudent] = field(default_factory=list)


@dataclass
class ParsedWorkbook:
    source_file: str
    room_number: str
    group_sheets: list[ParsedGroupSheet] = field(default_factory=list)
    catalog: list[ParsedLabWork] = field(default_factory=list)


def norm(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def room_number_for_file(path: Path) -> str:
    name = path.name
    for pattern, room in FILE_ROOM_BY_PATTERN:
        if pattern.search(name):
            return room
    raise ValueError(f"Не удалось определить аудиторию для файла {path.name}")


def is_meta_text(value: str) -> bool:
    lowered = value.lower()
    return any(
        token in lowered
        for token in (
            "группа",
            "руководитель",
            "куратор",
            "comment",
            "план",
            "факт",
            "ч-час",
            "время выполнения",
        )
    )


def looks_like_person_name(value: str) -> bool:
    parts = value.split()
    if len(parts) < 2 or len(parts) > 4:
        return False
    if not re.match(r"^[А-ЯA-ZЁ]", parts[0]):
        return False
    return all(re.match(r"^[А-ЯA-ZЁ][а-яa-zё-]+$", part) for part in parts)


def looks_like_lab_title(value: str) -> bool:
    if not value or is_meta_text(value) or looks_like_person_name(value):
        return False
    if value.startswith(("Лаб. раб", '"Лаб', "«Лаб")):
        return True
    if value.startswith(LAB_TITLE_PREFIXES):
        return True
    return len(value) >= 35


def looks_like_discipline(value: str) -> bool:
    if not value or is_meta_text(value) or looks_like_person_name(value):
        return False
    if re.fullmatch(r"\d+", value) or value == "Comment":
        return False
    return len(value) >= 8


def _header_discipline_cells(row) -> int:
    return sum(
        1
        for column_index, cell in enumerate(row)
        if column_index >= 2 and looks_like_discipline(norm(cell))
    )


def _header_lab_cells(row) -> int:
    return sum(
        1
        for column_index, cell in enumerate(row)
        if column_index >= 2 and looks_like_lab_title(norm(cell))
    )


def _data_cells(row) -> int:
    return sum(
        1
        for column_index, cell in enumerate(row)
        if column_index >= 2
        and (value := norm(cell))
        and not is_meta_text(value)
        and not looks_like_person_name(value)
        and not re.fullmatch(r"\d+", value)
        and value != "Comment"
    )


def detect_header_layout(rows) -> tuple[int, int]:
    for discipline_row_idx in range(min(6, len(rows))):
        if _header_discipline_cells(rows[discipline_row_idx]) == 0:
            continue
        for lab_row_idx in range(discipline_row_idx + 1, min(discipline_row_idx + 3, len(rows))):
            if _data_cells(rows[lab_row_idx]) > _data_cells(rows[discipline_row_idx]):
                return discipline_row_idx, lab_row_idx
    return 1, 2


def discipline_for_column(column: int, discipline_columns: dict[int, str]) -> str:
    candidates = [col for col in sorted(discipline_columns) if col <= column]
    if not candidates:
        return ""
    return discipline_columns[candidates[-1]]


def parse_student_name(full_name: str) -> tuple[str, str] | None:
    parts = full_name.split()
    if len(parts) < 2:
        return None
    if not re.match(r"^[А-ЯA-ZЁ]", parts[0]):
        return None
    last_name = parts[0]
    first_name = " ".join(parts[1:3]) if len(parts) >= 3 else parts[1]
    return last_name, first_name


def parse_group_sheet(worksheet) -> ParsedGroupSheet:
    rows = list(worksheet.iter_rows(max_row=140, values_only=True))
    durations: dict[int, int] = {}
    discipline_columns: dict[int, str] = {}
    lab_columns: dict[int, str] = {}

    if len(rows) < 3:
        return ParsedGroupSheet(name=worksheet.title.strip())

    discipline_row_idx, lab_row_idx = detect_header_layout(rows)

    for column_index, cell in enumerate(rows[discipline_row_idx]):
        if column_index < 2:
            continue
        value = norm(cell)
        if looks_like_discipline(value):
            discipline_columns[column_index] = value

    for column_index, cell in enumerate(rows[lab_row_idx]):
        if column_index < 2:
            continue
        value = norm(cell)
        if not value or is_meta_text(value) or value == "Comment":
            continue
        if re.fullmatch(r"\d+", value) or looks_like_person_name(value):
            continue
        lab_columns[column_index] = value

    for row in rows:
        if norm(row[2] if len(row) > 2 else "") == "Время выполнения, мин":
            for column_index, cell in enumerate(row):
                if column_index >= 3 and isinstance(cell, (int, float)) and cell:
                    durations[column_index] = int(cell)

    lab_works = [
        ParsedLabWork(
            discipline=discipline_for_column(column_index, discipline_columns),
            title=title,
            duration_minutes=durations.get(column_index),
        )
        for column_index, title in sorted(lab_columns.items())
    ]

    students: list[ParsedStudent] = []
    seen_numbers: set[int] = set()
    for row in rows:
        if len(row) < 3:
            continue
        number_cell = row[1] if isinstance(row[1], (int, float)) else row[0]
        full_name = norm(row[2] if isinstance(row[1], (int, float)) else row[1])
        if isinstance(row[0], (int, float)) and not isinstance(row[1], (int, float)):
            number_cell = row[0]
            full_name = norm(row[1])
        if not isinstance(number_cell, (int, float)):
            continue
        number = int(number_cell)
        if number in seen_numbers:
            continue
        parsed_name = parse_student_name(full_name)
        if not parsed_name:
            continue
        seen_numbers.add(number)
        last_name, first_name = parsed_name
        students.append(ParsedStudent(number=number, last_name=last_name, first_name=first_name))

    return ParsedGroupSheet(name=worksheet.title.strip(), lab_works=lab_works, students=students)


def parse_catalog_sheet(worksheet) -> list[ParsedLabWork]:
    catalog: list[ParsedLabWork] = []
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        number, stand, discipline, title = (row + (None,) * 6)[:4]
        title = norm(title)
        if not title:
            continue
        catalog_number = int(number) if isinstance(number, (int, float)) else None
        catalog.append(
            ParsedLabWork(
                discipline=norm(discipline),
                title=title,
                catalog_number=catalog_number,
                stand_name=norm(stand),
            )
        )
    return catalog


def parse_workbook(path: str | Path) -> ParsedWorkbook:
    file_path = Path(path)
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    parsed = ParsedWorkbook(source_file=file_path.name, room_number=room_number_for_file(file_path))

    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            if isinstance(worksheet, Chartsheet):
                continue
            if sheet_name == "Перечень лр":
                parsed.catalog = parse_catalog_sheet(worksheet)
                continue
            if sheet_name in SKIP_SHEETS:
                continue
            group_sheet = parse_group_sheet(worksheet)
            if group_sheet.lab_works or group_sheet.students:
                parsed.group_sheets.append(group_sheet)
    finally:
        workbook.close()

    return parsed
